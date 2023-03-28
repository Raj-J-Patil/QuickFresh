import openpyxl


class Excel:
    def get_max_row(self, path, sheetName):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheetName]
        num = sheet.max_row
        return num

    def get_max_column(self, path, sheetName):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheetName]
        num = sheet.max_column
        return num

    def get_cell_value(self, path, sheetName, row, column):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheetName]
        value = sheet.cell(row, column).value
        return value
